from openpyxl import load_workbook

wb = load_workbook("bikers.xlsx")

sheet = wb["bike_buyers"] 


headers = {}
for col in range(1, sheet.max_column + 1):
    headers[sheet.cell(row=1, column=col).value] = col

bike_yes = []
bike_no = []
region_counts = {}
gender_purchase = {"M": [0, 0], "F": [0, 0]}  


for row in range(2, sheet.max_row + 1):
    purchased = sheet.cell(row=row, column=headers["Purchased Bike"]).value
    income = sheet.cell(row=row, column=headers["Income"]).value
    region = sheet.cell(row=row, column=headers["Region"]).value
    gender = sheet.cell(row=row, column=headers["Gender"]).value

    # Clean income
    if isinstance(income, str):
        income = income.replace("$", "").replace(",", "")
    income = float(income)

    # Bike purchasers
    if purchased == "Yes":
        bike_yes.append(income)
    elif purchased == "No":
        bike_no.append(income)

    # Count region
    if region in region_counts:
        region_counts[region][0] += 1
    else:
        region_counts[region] = [1, 0]

    # Region and purchase status
    if purchased == "Yes":
        region_counts[region][1] += 1

    # Gender-based purchase 
    if gender in gender_purchase:
        gender_purchase[gender][1] += 1  
        if purchased == "Yes":
            gender_purchase[gender][0] += 1  

# Print average incomes
if len(bike_yes) > 0:
    avg_income_yes = sum(bike_yes) // len(bike_yes)
    print("The average income of people who purchased a bike is:", avg_income_yes)

if len(bike_no) > 0:
    avg_income_no = sum(bike_no) // len(bike_no)
    print("The average income of people who didn't purchase a bike is:", avg_income_no)
print()

# Region counts
print("Regions represented here:")
for region, counts in region_counts.items():
    print(region, ":", counts[0])

print()

# Region-wise purchase percentage
print("Percentage of citizens buying bikes in each region:")
for region, counts in region_counts.items():
    total = counts[0]
    bought = counts[1]
    percentage = (bought / total) * 100 if total > 0 else 0
    print(region, ":", round(percentage, 2), "%")

print()

# Gender-based purchase percentages
print("Gender Based Purchase of Bikes")
print("Female buying Bike:", round((gender_purchase["F"][0] / gender_purchase["F"][1]) * 100, 2))
print("Male buying Bike:", round((gender_purchase["M"][0] / gender_purchase["M"][1]) * 100, 2))
