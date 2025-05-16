import pandas as pd
from openpyxl import Workbook

def read_file(filename):
    rows = []
    year = filename.replace(".txt", "").replace("Elections", "")

    file = open(filename, "r")
    for line in file:
        parts = line.strip().split()
        

        try:
            constituency = parts[1]

            name1 = parts[2]
            gender1 = parts[3]
            party1 = parts[4]
            votes1 = int(parts[5])

            name2 = parts[6]
            gender2 = parts[7]
            party2 = parts[8]
            votes2 = int(parts[9])

            row1 = [year, constituency, name1, gender1, party1, votes1]
            row2 = [year, constituency, name2, gender2, party2, votes2]

            rows.append(row1)
            rows.append(row2)
        except:
            continue

    file.close()
    return year, rows

def analyze_data(all_rows):
    df = pd.DataFrame(all_rows, columns=["Year", "Constituency", "Candidate", "Gender", "Party", "Votes"])
    df["Votes"] = pd.to_numeric(df["Votes"], errors="coerce")
    df["Winner"] = df.groupby(["Year", "Constituency"])["Votes"].transform("max") == df["Votes"]

    # Count seats won by party
    seats_won = df[df["Winner"] == True]["Party"].value_counts()

    # Find highest winning margin
    margin_list = []
    grouped = df.groupby(["Year", "Constituency"])
    for (year, constituency), group in grouped:
        sorted_group = group.sort_values("Votes", ascending=False)
        if len(sorted_group) >= 2:
            winner = sorted_group.iloc[0]
            runner_up = sorted_group.iloc[1]
            margin = winner["Votes"] - runner_up["Votes"]
            margin_row = [
                winner["Year"], winner["Constituency"], winner["Candidate"],
                winner["Party"], winner["Votes"], margin
            ]
            margin_list.append(margin_row)

    female_all = df[df["Gender"].str.upper() == "F"]
    female_winners = female_all[female_all["Winner"] == True]

    return df, seats_won, margin_list, female_all, female_winners

# Save all output to Excel
def save_to_excel(yearly_data, seats_won, margin_list, female_all, female_winners):
    wb = Workbook()
    wb.remove(wb["Sheet"])  

    
    for year, rows in yearly_data.items():
        ws = wb.create_sheet(title=f"Elections{year}")
        ws.append(["Year", "Constituency", "Candidate", "Gender", "Party", "Votes"])
        for row in rows:
            ws.append(row)

    
    ws1 = wb.create_sheet("Seats Won")
    ws1.append(["Party", "Seats Won"])
    for party, count in seats_won.items():
        ws1.append([party, count])

    
    ws2 = wb.create_sheet("Highest Margin Winner")
    ws2.append(["Year", "Constituency", "Candidate", "Party", "Votes", "Margin"])
    if margin_list:
        ws2.append(margin_list[0])

    ws3 = wb.create_sheet("Female Candidates")
    ws3.append(["Total Female Contested", "Total Female Winners"])
    ws3.append([len(female_all), len(female_winners)])

    wb.save("Election_Results_Analysis.xlsx")
    print("\n")
    print("Excel file 'Election_Results_Analysis.xlsx' created.")

# calling files
files = ["Elections2004.txt", "Elections2009.txt", "Elections2014.txt"]
all_rows = []
yearly_data = {}

for file in files:
    year, rows = read_file(file)
    yearly_data[year] = rows
    all_rows.extend(rows)

df, seats_won, margin_list, female_all, female_winners = analyze_data(all_rows)
print("Seats won by each party:")
for party, count in seats_won.items():
    print(f"{party}: {count}")

print("Highest margin winner:")
if margin_list:
    top = margin_list[0]
    print(f"{top[2]} ({top[3]}) from {top[1]} in {top[0]} - Margin: {top[5]} votes")

print("Female candidates:")
print("Total female candidates:", len(female_all))
print("Total female winners:", len(female_winners))

# Save to Excel
save_to_excel(yearly_data, seats_won, margin_list, female_all, female_winners)
